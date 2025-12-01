#!/usr/bin/env python3
"""
Excel Writer for Enriched Reports

Writes enriched Webgains records to Excel file.
"""

import logging
from typing import List
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from scripts.generate_report_from_webgains.models import EnrichedRecord

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Writes enriched records to Excel file"""

    # Define original Webgains columns (Sheet 1)
    ORIGINAL_COLUMNS = [
        ("Affiliate", "affiliate"),
        ("Sale", "sale"),
        ("Commission", "commission"),
        ("Override", "override"),
        ("Date & Time", "date_time"),
        ("Order Reference", "order_reference"),
        ("Country", "country"),
        ("Commission Type", "commission_type"),
        ("%", "percentage"),
    ]

    # Define enriched data columns (Sheet 2 - includes original + enriched)
    ENRICHED_COLUMNS = [
        # Original Webgains columns
        ("Affiliate", "affiliate"),
        ("Sale", "sale"),
        ("Commission", "commission"),
        ("Override", "override"),
        ("Date & Time", "date_time"),
        ("Order Reference", "order_reference"),
        ("Country", "country"),
        ("Commission Type", "commission_type"),
        ("%", "percentage"),

        # Enriched Shopify columns
        ("Product Names", "product_names"),
        ("Variant Names", "variant_names"),
        ("Product SKUs", "product_skus"),
        ("Financial Status", "financial_status"),
        ("Fulfillment Status", "fulfillment_status"),
        ("Is Cancelled", "is_cancelled"),
        ("Order Status Notes", "order_status_notes"),
        ("Refund Amount", "refund_amount"),  # New column for partial refund amount
        ("Return Status", "return_status"),  # New column for return in process
        ("Issue Type", "issue_type"),  # New column for issue type
        ("Customer Email", "customer_email"),
        ("Customer First Name", "customer_first_name"),
        ("Customer Last Name", "customer_last_name"),
        ("Customer Phone", "customer_phone"),
        ("Customer Type", "customer_type"),
        ("Order Number for Customer", "order_number_for_customer"),
        ("First Visit Source", "first_visit_source"),
        ("UTM Source", "utm_source"),
        ("UTM Medium", "utm_medium"),
        ("UTM Campaign", "utm_campaign"),
        ("UTM Term", "utm_term"),
        ("UTM Content", "utm_content"),
        ("Error", "error_message")
    ]

    def __init__(self, output_path: str):
        """
        Initialize Excel writer

        Args:
            output_path: Path for output Excel file
        """
        self.output_path = Path(output_path)
        self.workbook = None
        self.original_sheet = None
        self.enriched_sheet = None

    def write_records(self, records: List[EnrichedRecord]) -> bool:
        """
        Write enriched records to Excel file with two sheets

        Args:
            records: List of enriched records to write

        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Writing {len(records)} enriched records to {self.output_path}")

            # Create workbook
            self.workbook = openpyxl.Workbook()

            # Create Sheet 1: Analysis Summary
            self.analysis_sheet = self.workbook.active
            self.analysis_sheet.title = "Analysis Summary"
            self._write_analysis_sheet(self.analysis_sheet, records)

            # Create Sheet 2: Enriched Data (Newcop Data)
            self.enriched_sheet = self.workbook.create_sheet(title="Newcop Data")
            self._write_sheet(self.enriched_sheet, records, self.ENRICHED_COLUMNS)
            self._apply_conditional_formatting(self.enriched_sheet, records)

            # Create Sheet 3: Original Webgains Data
            self.original_sheet = self.workbook.create_sheet(title="Original Webgains Data")
            self._write_sheet(self.original_sheet, records, self.ORIGINAL_COLUMNS)

            # Save workbook
            self.workbook.save(self.output_path)
            logger.info(f"Successfully wrote enriched report to {self.output_path}")

            return True

        except Exception as e:
            logger.error(f"Error writing Excel file: {e}")
            return False

    def _write_sheet(self, worksheet, records: List[EnrichedRecord], columns):
        """
        Write data to a specific sheet

        Args:
            worksheet: Worksheet to write to
            records: List of enriched records
            columns: List of column definitions
        """
        # Write header row
        self._write_header(worksheet, columns)

        # Write data rows
        self._write_data_rows(worksheet, records, columns)

        # Apply formatting
        self._apply_formatting(worksheet, columns)

    def _write_header(self, worksheet, columns):
        """Write header row with column names"""
        for col_idx, (header_name, _) in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = header_name

            # Style header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _write_data_rows(self, worksheet, records: List[EnrichedRecord], columns):
        """
        Write data rows

        Args:
            worksheet: Worksheet to write to
            records: List of enriched records
            columns: List of column definitions
        """
        for row_idx, record in enumerate(records, start=2):
            for col_idx, (_, field_name) in enumerate(columns, start=1):
                # Get value from record
                value = self._get_field_value(record, field_name)

                # Write cell value
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Format numeric cells
                if field_name in ["sale", "commission", "override"]:
                    if isinstance(value, (int, float)):
                        cell.number_format = '€#,##0.00'
                elif field_name == "percentage":
                    # Force percentage to be treated as text to preserve "5%" format
                    cell.number_format = '@'  # @ means text format in Excel

    def _get_field_value(self, record: EnrichedRecord, field_name: str):
        """
        Get field value from enriched record

        Args:
            record: EnrichedRecord object
            field_name: Field name to retrieve

        Returns:
            Field value
        """
        # Direct attributes
        if hasattr(record, field_name):
            value = getattr(record, field_name)
            return value if value is not None else ""

        # Properties (computed fields)
        try:
            value = getattr(record, field_name)
            return value if value is not None else ""
        except AttributeError:
            return ""

    def _apply_formatting(self, worksheet, columns):
        """Apply formatting to worksheet"""
        # Auto-size columns
        for col_idx in range(1, len(columns) + 1):
            column_letter = get_column_letter(col_idx)

            # Calculate max width for column
            max_width = 10
            for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        max_width = max(max_width, cell_len)

            # Set column width (cap at 50)
            worksheet.column_dimensions[column_letter].width = min(max_width + 2, 50)

        # Freeze header row
        worksheet.freeze_panes = "A2"

        # Add alternating row colors for readability
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_idx in range(3, worksheet.max_row + 1, 2):  # Every other row starting from row 3
            for col_idx in range(1, len(columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = light_fill

    def _write_analysis_sheet(self, worksheet, records: List[EnrichedRecord]):
        """
        Write analysis summary with statistics

        Args:
            worksheet: Worksheet to write to
            records: List of enriched records
        """
        from collections import Counter

        row_idx = 1

        # Title
        title_cell = worksheet.cell(row=row_idx, column=1)
        title_cell.value = "Webgains Report Analysis Summary"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        worksheet.merge_cells(f'A{row_idx}:D{row_idx}')
        row_idx += 2

        # 1. Country Distribution Analysis
        country_stats = self._calculate_country_stats(records)
        row_idx = self._write_stat_section(worksheet, row_idx, "Country Distribution", country_stats)
        row_idx += 1

        # 2. Commission Type Analysis (Retail vs Resell)
        commission_type_stats = self._calculate_commission_type_stats(records)
        row_idx = self._write_stat_section(worksheet, row_idx, "Commission Type Distribution", commission_type_stats)
        row_idx += 1

        # 3. Customer Type Analysis (New vs Repeat)
        customer_type_stats = self._calculate_customer_type_stats(records)
        row_idx = self._write_stat_section(worksheet, row_idx, "Customer Type Distribution", customer_type_stats)
        row_idx += 1

        # 4. Product Distribution Analysis
        product_stats = self._calculate_product_stats(records)
        row_idx = self._write_stat_section(worksheet, row_idx, "Top Products Distribution", product_stats)
        row_idx += 1

        # 5. Orders Requiring Review (with hyperlinks to Newcop Data sheet)
        review_orders = self._calculate_review_orders(records)
        row_idx = self._write_review_orders_section(worksheet, row_idx, "Orders Requiring Review", review_orders)

        # Format columns
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 25
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 20
        worksheet.column_dimensions['F'].width = 20

    def _calculate_country_stats(self, records: List[EnrichedRecord]) -> List[tuple]:
        """Calculate country distribution statistics"""
        from collections import Counter

        country_counts = Counter()
        total = 0

        for record in records:
            if record.country:
                country_counts[record.country] += 1
                total += 1

        # Sort by count descending
        stats = []
        for country, count in country_counts.most_common():
            percentage = (count / total * 100) if total > 0 else 0
            stats.append((country, count, f"{percentage:.2f}%"))

        return stats

    def _calculate_commission_type_stats(self, records: List[EnrichedRecord]) -> List[tuple]:
        """Calculate commission type (retail vs resell) statistics"""
        from collections import Counter

        commission_counts = Counter()
        total = 0

        for record in records:
            if record.commission_type:
                # Normalize to lowercase for comparison
                comm_type = record.commission_type.strip().lower()
                commission_counts[comm_type] += 1
                total += 1

        # Sort by count descending
        stats = []
        for comm_type, count in commission_counts.most_common():
            percentage = (count / total * 100) if total > 0 else 0
            # Capitalize first letter for display
            display_name = comm_type.capitalize()
            stats.append((display_name, count, f"{percentage:.2f}%"))

        return stats

    def _calculate_customer_type_stats(self, records: List[EnrichedRecord]) -> List[tuple]:
        """Calculate customer type (new vs repeat) statistics"""
        new_customers = 0
        repeat_customers = 0
        unknown = 0

        for record in records:
            customer_type = record.customer_type
            if "First Order" in customer_type:
                new_customers += 1
            elif "Repeat Customer" in customer_type:
                repeat_customers += 1
            else:
                unknown += 1

        total = len(records)
        stats = []

        if new_customers > 0:
            percentage = (new_customers / total * 100) if total > 0 else 0
            stats.append(("New Customers (First Order)", new_customers, f"{percentage:.2f}%"))

        if repeat_customers > 0:
            percentage = (repeat_customers / total * 100) if total > 0 else 0
            stats.append(("Repeat Customers", repeat_customers, f"{percentage:.2f}%"))

        if unknown > 0:
            percentage = (unknown / total * 100) if total > 0 else 0
            stats.append(("Unknown", unknown, f"{percentage:.2f}%"))

        return stats

    def _calculate_product_stats(self, records: List[EnrichedRecord]) -> List[tuple]:
        """Calculate product distribution statistics (by product name, not variant)"""
        from collections import Counter

        product_counts = Counter()
        total = 0

        for record in records:
            # Get product names (comma-separated if multiple)
            product_names_str = record.product_names
            if product_names_str:
                # Split by comma and process each product
                products = [p.strip() for p in product_names_str.split(',')]
                for product in products:
                    if product:
                        product_counts[product] += 1
                        total += 1

        # Get top 20 products
        stats = []
        for product, count in product_counts.most_common(20):
            percentage = (count / total * 100) if total > 0 else 0
            stats.append((product, count, f"{percentage:.2f}%"))

        return stats

    def _calculate_review_orders(self, records: List[EnrichedRecord]) -> List[tuple]:
        """
        Calculate orders that require review (yellow or red rows)
        Returns list of (row_number, order_reference, issue_type, financial_status, fulfillment_status, refund_amount, return_status, severity)
        severity: 'urgent' for red rows, 'warning' for yellow rows
        """
        review_orders = []

        for idx, record in enumerate(records, start=2):  # Start at row 2 (after header)
            financial_status = record.financial_status.upper() if record.financial_status else ""
            fulfillment_status = record.fulfillment_status.upper() if record.fulfillment_status else ""
            refund_amount = record.refund_amount  # Get refund amount from property
            return_status = record.return_status  # Get return status from property
            issue_type = record.issue_type  # Get issue type directly from Newcop Data

            severity = ""

            # Check if this order needs review
            needs_review = False

            # Check fulfillment status first (higher priority - red)
            if fulfillment_status and fulfillment_status == "UNFULFILLED":
                severity = "urgent"
                needs_review = True
            # Check financial status (yellow)
            elif financial_status and financial_status != "PAID":
                severity = "warning"
                needs_review = True

            if needs_review:
                order_ref = record.order_reference if record.order_reference else "N/A"
                review_orders.append((
                    idx,  # Row number in Newcop Data sheet
                    order_ref,
                    issue_type if issue_type else "",  # Use issue_type from Newcop Data
                    financial_status if financial_status else "N/A",
                    fulfillment_status if fulfillment_status else "N/A",
                    refund_amount if refund_amount else "",  # Include refund amount
                    return_status if return_status else "",  # Include return status
                    severity
                ))

        return review_orders

    def _write_review_orders_section(self, worksheet, start_row: int, title: str, review_orders: List[tuple]) -> int:
        """
        Write review orders section with hyperlinks to Newcop Data sheet

        Args:
            worksheet: Worksheet to write to
            start_row: Starting row number
            title: Section title
            review_orders: List of (row_number, order_reference, issue_type, financial_status, fulfillment_status, refund_amount, return_status, severity) tuples

        Returns:
            Next available row number
        """
        row_idx = start_row

        # Section title
        title_cell = worksheet.cell(row=row_idx, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        worksheet.merge_cells(f'A{row_idx}:F{row_idx}')  # Extended to F to include return status column
        row_idx += 1

        # Summary counts
        urgent_count = sum(1 for _, _, _, _, _, _, _, severity in review_orders if severity == "urgent")
        warning_count = sum(1 for _, _, _, _, _, _, _, severity in review_orders if severity == "warning")

        summary_cell = worksheet.cell(row=row_idx, column=1)
        summary_cell.value = f"🔴 Urgent: {urgent_count}  |  🟡 Warning: {warning_count}  |  Total: {len(review_orders)}"
        summary_cell.font = Font(bold=True, size=11)
        worksheet.merge_cells(f'A{row_idx}:F{row_idx}')  # Extended to F
        row_idx += 1

        if not review_orders:
            # No issues found
            no_issues_cell = worksheet.cell(row=row_idx, column=1)
            no_issues_cell.value = "✅ All orders are in good standing - no issues found!"
            no_issues_cell.font = Font(color="008000", bold=True)  # Green text
            worksheet.merge_cells(f'A{row_idx}:F{row_idx}')  # Extended to F
            row_idx += 1
            return row_idx

        # Header row
        headers = ["Order Reference", "Issue Type", "Financial Status", "Fulfillment Status", "Refund Amount", "Return Status"]
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_idx += 1

        # Data rows with hyperlinks
        for newcop_row_num, order_ref, issue_type, financial_status, fulfillment_status, refund_amount, return_status, severity in review_orders:
            # Column 1: Order Reference with hyperlink
            cell = worksheet.cell(row=row_idx, column=1)
            cell.value = order_ref
            # Create hyperlink to Newcop Data sheet
            cell.hyperlink = f"#'Newcop Data'!A{newcop_row_num}"
            cell.font = Font(color="0000FF", underline="single")  # Blue underlined link
            cell.alignment = Alignment(horizontal="left")

            # Column 2: Issue Type
            cell = worksheet.cell(row=row_idx, column=2)
            cell.value = issue_type
            # Color code by severity
            if severity == "urgent":
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
                cell.font = Font(color="FFFFFF", bold=True)  # White text
            else:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

            # Column 3: Financial Status
            cell = worksheet.cell(row=row_idx, column=3)
            cell.value = financial_status
            cell.alignment = Alignment(horizontal="center")

            # Column 4: Fulfillment Status
            cell = worksheet.cell(row=row_idx, column=4)
            cell.value = fulfillment_status
            cell.alignment = Alignment(horizontal="center")

            # Column 5: Refund Amount
            cell = worksheet.cell(row=row_idx, column=5)
            cell.value = refund_amount
            cell.alignment = Alignment(horizontal="center")
            # Bold if there's a refund amount
            if refund_amount:
                cell.font = Font(bold=True, color="FF0000")  # Bold red for refund amounts

            # Column 6: Return Status
            cell = worksheet.cell(row=row_idx, column=6)
            cell.value = return_status
            cell.alignment = Alignment(horizontal="center")
            # Bold orange if there's a return in process
            if return_status:
                cell.font = Font(bold=True, color="FF6600")  # Bold orange for return status

            row_idx += 1

        return row_idx

    def _write_stat_section(self, worksheet, start_row: int, title: str, stats: List[tuple]) -> int:
        """
        Write a statistics section

        Args:
            worksheet: Worksheet to write to
            start_row: Starting row number
            title: Section title
            stats: List of (name, count, percentage) tuples

        Returns:
            Next available row number
        """
        row_idx = start_row

        # Section title
        title_cell = worksheet.cell(row=row_idx, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        worksheet.merge_cells(f'A{row_idx}:D{row_idx}')
        row_idx += 1

        # Header row
        headers = ["Category", "Count", "Percentage"]
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_idx += 1

        # Data rows
        for name, count, percentage in stats:
            worksheet.cell(row=row_idx, column=1).value = name
            worksheet.cell(row=row_idx, column=2).value = count
            worksheet.cell(row=row_idx, column=2).alignment = Alignment(horizontal="right")
            worksheet.cell(row=row_idx, column=3).value = percentage
            worksheet.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right")
            row_idx += 1

        # Total row (if applicable)
        if stats:
            total_count = sum(count for _, count, _ in stats)
            total_cell = worksheet.cell(row=row_idx, column=1)
            total_cell.value = "Total"
            total_cell.font = Font(bold=True)
            count_cell = worksheet.cell(row=row_idx, column=2)
            count_cell.value = total_count
            count_cell.font = Font(bold=True)
            count_cell.alignment = Alignment(horizontal="right")
            pct_cell = worksheet.cell(row=row_idx, column=3)
            pct_cell.value = "100.00%"
            pct_cell.font = Font(bold=True)
            pct_cell.alignment = Alignment(horizontal="right")
            row_idx += 1

        return row_idx

    def _apply_conditional_formatting(self, worksheet, records: List[EnrichedRecord]):
        """
        Apply conditional formatting to Newcop Data sheet
        - Yellow background if Financial Status is not PAID
        - Red background if Fulfillment Status is UNFULFILLED (overrides yellow)

        Args:
            worksheet: Worksheet to apply formatting to
            records: List of enriched records
        """
        # Find column indices for Financial Status and Fulfillment Status
        financial_status_col = None
        fulfillment_status_col = None

        for col_idx, (header_name, field_name) in enumerate(self.ENRICHED_COLUMNS, start=1):
            if field_name == "financial_status":
                financial_status_col = col_idx
            elif field_name == "fulfillment_status":
                fulfillment_status_col = col_idx

        if not financial_status_col or not fulfillment_status_col:
            logger.warning("Could not find Financial Status or Fulfillment Status columns for conditional formatting")
            return

        # Define fill colors
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
        white_font = Font(color="FFFFFF")  # White text for red background

        # Apply formatting to data rows (starting from row 2, after header)
        for row_idx, record in enumerate(records, start=2):
            financial_status = record.financial_status.upper() if record.financial_status else ""
            fulfillment_status = record.fulfillment_status.upper() if record.fulfillment_status else ""

            # Determine row color
            row_fill = None
            apply_white_font = False

            # First check financial status (yellow for non-PAID)
            if financial_status and financial_status != "PAID":
                row_fill = yellow_fill

            # Then check fulfillment status (red for UNFULFILLED, overrides yellow)
            if fulfillment_status and fulfillment_status == "UNFULFILLED":
                row_fill = red_fill
                apply_white_font = True

            # Apply formatting to entire row if needed
            if row_fill:
                for col_idx in range(1, len(self.ENRICHED_COLUMNS) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = row_fill
                    if apply_white_font:
                        cell.font = white_font

        logger.info(f"Applied conditional formatting to Newcop Data sheet")

    def close(self):
        """Close workbook"""
        if self.workbook:
            self.workbook.close()
            logger.debug("Workbook closed")
