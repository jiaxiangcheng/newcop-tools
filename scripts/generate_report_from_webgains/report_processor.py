#!/usr/bin/env python3
"""
Report Processor for Webgains Excel Files

Reads and parses Webgains sales reports from Excel files.
"""

import logging
from typing import List, Optional
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from scripts.generate_report_from_webgains.models import WebgainsRecord

logger = logging.getLogger(__name__)


class ReportProcessor:
    """Processes Webgains Excel reports"""

    def __init__(self, file_path: str):
        """
        Initialize report processor

        Args:
            file_path: Path to Webgains Excel file
        """
        self.file_path = Path(file_path)
        self.workbook = None
        self.worksheet: Optional[Worksheet] = None

    def load_workbook(self) -> bool:
        """
        Load Excel workbook

        Returns:
            True if successful, False otherwise
        """
        try:
            if not self.file_path.exists():
                logger.error(f"File not found: {self.file_path}")
                return False

            logger.info(f"Loading Excel file: {self.file_path}")
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            self.worksheet = self.workbook.active

            logger.info(f"Loaded workbook with {self.worksheet.max_row} rows")
            return True

        except Exception as e:
            logger.error(f"Error loading workbook: {e}")
            return False

    def parse_records(self, limit: Optional[int] = None) -> List[WebgainsRecord]:
        """
        Parse Webgains records from Excel

        Args:
            limit: Optional limit on number of records to parse

        Returns:
            List of WebgainsRecord objects
        """
        if not self.worksheet:
            logger.error("Workbook not loaded")
            return []

        records = []

        try:
            # Find header row (typically row 1)
            header_row = self._find_header_row()
            if not header_row:
                logger.error("Could not find header row in Excel file")
                return []

            # Map column names to indices
            column_map = self._build_column_map(header_row)
            logger.info(f"Column mapping: {column_map}")

            # Parse data rows
            max_row = self.worksheet.max_row
            if limit:
                max_row = min(max_row, header_row + limit)

            logger.info(f"Parsing rows {header_row + 1} to {max_row}")

            for row_idx in range(header_row + 1, max_row + 1):
                row_values = [cell.value for cell in self.worksheet[row_idx]]

                # Skip empty rows
                if all(v is None or str(v).strip() == "" for v in row_values):
                    continue

                try:
                    record = self._parse_row(row_values, column_map)
                    if record:
                        records.append(record)
                except Exception as e:
                    logger.warning(f"Error parsing row {row_idx}: {e}")
                    continue

            logger.info(f"Successfully parsed {len(records)} records")
            return records

        except Exception as e:
            logger.error(f"Error parsing records: {e}")
            return []

    def _find_header_row(self) -> Optional[int]:
        """
        Find the header row in the worksheet

        Returns:
            Row index (1-based) or None if not found
        """
        # Look for header row in first 5 rows
        for row_idx in range(1, min(6, self.worksheet.max_row + 1)):
            row_values = [str(cell.value).lower() if cell.value else "" for cell in self.worksheet[row_idx]]

            # Check if this looks like a header row
            if any("affiliate" in val or "order" in val or "commission" in val for val in row_values):
                logger.info(f"Found header row at index {row_idx}")
                return row_idx

        return None

    def _build_column_map(self, header_row: int) -> dict:
        """
        Build mapping of column names to indices

        Args:
            header_row: Row index of header row

        Returns:
            Dictionary mapping normalized column names to column indices
        """
        column_map = {}

        for col_idx, cell in enumerate(self.worksheet[header_row]):
            if cell.value:
                # Normalize column name
                col_name = str(cell.value).strip().lower()

                # Map to standard field names
                if "affiliate" in col_name:
                    column_map["affiliate"] = col_idx
                elif "sale" in col_name and "commission" not in col_name:
                    column_map["sale"] = col_idx
                elif "commission" in col_name and "type" not in col_name:
                    column_map["commission"] = col_idx
                elif "override" in col_name:
                    column_map["override"] = col_idx
                elif "date" in col_name and "time" in col_name:
                    column_map["date_time"] = col_idx
                elif "order" in col_name and "reference" in col_name:
                    column_map["order_reference"] = col_idx
                elif "country" in col_name:
                    column_map["country"] = col_idx
                elif "commission" in col_name and "type" in col_name:
                    column_map["commission_type"] = col_idx
                elif col_name == "%" or "percentage" in col_name:
                    column_map["percentage"] = col_idx

        return column_map

    def _parse_row(self, row_values: List, column_map: dict) -> Optional[WebgainsRecord]:
        """
        Parse a single row into WebgainsRecord

        Args:
            row_values: List of cell values
            column_map: Column name to index mapping

        Returns:
            WebgainsRecord or None if parsing fails
        """
        try:
            # Extract values using column map
            record_data = {}

            for field_name, col_idx in column_map.items():
                if col_idx < len(row_values):
                    value = row_values[col_idx]

                    # Clean and convert value
                    if value is not None:
                        if field_name in ["sale", "commission", "override"]:
                            # Convert to float, removing currency symbols
                            if isinstance(value, (int, float)):
                                record_data[field_name] = float(value)
                            else:
                                # Try to parse string value
                                cleaned = str(value).replace("€", "").replace(",", "").strip()
                                try:
                                    record_data[field_name] = float(cleaned) if cleaned else None
                                except ValueError:
                                    record_data[field_name] = None
                        elif field_name == "percentage":
                            # Handle percentage - ensure it's formatted as string with %
                            if isinstance(value, (int, float)):
                                # If it's a decimal (0.05), convert to percentage string (5%)
                                if 0 <= value <= 1:
                                    record_data[field_name] = f"{int(value * 100)}%"
                                else:
                                    # If it's already a whole number (5), just add %
                                    record_data[field_name] = f"{int(value)}%"
                            else:
                                # If it's already a string, keep it but ensure % is present
                                value_str = str(value).strip()
                                if not value_str.endswith("%"):
                                    # Try to parse and add %
                                    try:
                                        num_val = float(value_str.replace("%", ""))
                                        if 0 <= num_val <= 1:
                                            record_data[field_name] = f"{int(num_val * 100)}%"
                                        else:
                                            record_data[field_name] = f"{int(num_val)}%"
                                    except ValueError:
                                        record_data[field_name] = value_str
                                else:
                                    record_data[field_name] = value_str
                        elif field_name == "date_time":
                            # Keep as string or convert datetime to string
                            if hasattr(value, "strftime"):
                                record_data[field_name] = value.strftime("%m/%d/%y %H:%M")
                            else:
                                record_data[field_name] = str(value)
                        else:
                            # Keep as string
                            record_data[field_name] = str(value).strip()

            # Create WebgainsRecord
            return WebgainsRecord(**record_data)

        except Exception as e:
            logger.warning(f"Error creating record from row: {e}")
            return None

    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()
            logger.debug("Workbook closed")
