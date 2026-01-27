"""
Manager for setting product types based on collection rules.
"""
import logging
import time
from typing import List, Dict, Any, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import ProductTypeUpdate, TypeSyncResult

logger = logging.getLogger(__name__)


class ProductTypeManager:
    """Manages product type updates based on collection rules."""

    # Collection configurations
    COLLECTIONS = {
        "639759778133": {
            "name": "Accessories Collection",
            "type": "Accessories"
        },
        "639750963541": {
            "name": "Sneakers Collection",
            "type_rules": {
                "retail_tag": "Retail Sneakers",
                "no_retail_tag": "Resell Sneakers"
            }
        },
        "639759647061": {
            "name": "Clothing Collection",
            "type": "Clothing"
        }
    }

    MAX_CONCURRENT_UPDATES = 5

    def __init__(self, shopify_client):
        """
        Initialize the product type manager.

        Args:
            shopify_client: ShopifyClient instance
        """
        self.shopify_client = shopify_client

    def get_product_details(self, product_id: int) -> Optional[Dict[str, Any]]:
        """
        Get detailed product information including type and tags.

        Args:
            product_id: Shopify product ID

        Returns:
            Product details dictionary or None if not found
        """
        query = """
        query getProduct($id: ID!) {
          product(id: $id) {
            id
            legacyResourceId
            title
            productType
            tags
          }
        }
        """

        gid = f"gid://shopify/Product/{product_id}"
        variables = {"id": gid}

        try:
            response = self.shopify_client.execute_graphql(query, variables)

            if "errors" in response:
                error_messages = [err.get("message", "Unknown error") for err in response["errors"]]
                logger.warning(f"GraphQL errors for product {product_id}: {', '.join(error_messages)}")
                return None

            product = response.get("data", {}).get("product")
            if not product:
                logger.warning(f"Product {product_id} not found")
                return None

            # Handle tags - GraphQL returns as array
            tags = product.get("tags", [])
            if isinstance(tags, str):
                # If it's a string (comma-separated), convert to list
                tags = [tag.strip() for tag in tags.split(",") if tag.strip()]

            return {
                "id": product.get("legacyResourceId"),
                "gid": product.get("id"),
                "title": product.get("title"),
                "type": product.get("productType"),
                "tags": tags
            }

        except Exception as e:
            logger.error(f"Error fetching product {product_id} details: {e}")
            return None

    def determine_product_type(self, collection_id: str, product_tags: List[str]) -> Optional[str]:
        """
        Determine the product type based on collection and product tags.

        Args:
            collection_id: Collection ID
            product_tags: List of product tags

        Returns:
            Product type string or None if collection not configured
        """
        collection_config = self.COLLECTIONS.get(collection_id)
        if not collection_config:
            logger.warning(f"Collection {collection_id} not configured")
            return None

        # Accessories collection - always "Accessories"
        if collection_id == "639759778133":
            return "Accessories"

        # Clothing collection - always "Clothing"
        if collection_id == "639759647061":
            return "Clothing"

        # Sneakers collection - check for retail tag
        if collection_id == "639750963541":
            # Convert tags list to lowercase for case-insensitive matching
            tags_lower = [tag.lower() if isinstance(tag, str) else str(tag).lower() for tag in product_tags]
            
            # Check if "retail" tag exists
            if "retail" in tags_lower:
                return "Retail Sneakers"
            else:
                return "Resell Sneakers"

        return None

    def update_product_type(self, product_id: int, product_gid: str, new_type: str) -> bool:
        """
        Update product type using GraphQL.

        Args:
            product_id: Numeric product ID (for logging)
            product_gid: GraphQL Global ID (gid://shopify/Product/...)
            new_type: New product type value

        Returns:
            True if successful, False otherwise
        """
        mutation = """
        mutation updateProductType($input: ProductInput!) {
          productUpdate(input: $input) {
            product {
              id
              productType
            }
            userErrors {
              field
              message
            }
          }
        }
        """

        variables = {
            "input": {
                "id": product_gid,
                "productType": new_type
            }
        }

        try:
            response = self.shopify_client.execute_graphql(mutation, variables)

            # Check for errors
            if "errors" in response:
                error_messages = [err.get("message", "Unknown error") for err in response["errors"]]
                logger.error(f"GraphQL errors updating product {product_id} type: {', '.join(error_messages)}")
                return False

            # Check for user errors
            data = response.get("data", {})
            product_update = data.get("productUpdate", {})
            user_errors = product_update.get("userErrors", [])

            if user_errors:
                error_messages = [f"{err['field']}: {err['message']}" for err in user_errors]
                logger.error(f"User errors updating product {product_id} type: {', '.join(error_messages)}")
                return False

            updated_product = product_update.get("product")
            if updated_product:
                logger.debug(f"Successfully updated product {product_id} type to '{new_type}'")
                return True
            else:
                logger.warning(f"No product returned for product {product_id}")
                return False

        except Exception as e:
            logger.error(f"Exception updating product {product_id} type: {e}")
            return False

    def process_collection(self, collection_id: str, dry_run: bool = False) -> Dict[str, Any]:
        """
        Process a single collection and update product types.

        Args:
            collection_id: Collection ID to process
            dry_run: If True, only analyze without making changes

        Returns:
            Dictionary with processing results
        """
        collection_config = self.COLLECTIONS.get(collection_id)
        if not collection_config:
            logger.error(f"Collection {collection_id} not configured")
            return {
                "collection_id": collection_id,
                "success": False,
                "error": "Collection not configured",
                "total": 0,
                "updated": 0,
                "skipped": 0,
                "failed": 0
            }

        collection_name = collection_config.get("name", f"Collection {collection_id}")
        logger.info(f"📦 Processing collection: {collection_name} (ID: {collection_id})")

        # Get all products in collection
        try:
            products = self.shopify_client.get_collection_products(collection_id)
            logger.info(f"Found {len(products)} products in collection {collection_id}")
        except Exception as e:
            logger.error(f"Error fetching products from collection {collection_id}: {e}")
            return {
                "collection_id": collection_id,
                "success": False,
                "error": str(e),
                "total": 0,
                "updated": 0,
                "skipped": 0,
                "failed": 0
            }

        if not products:
            logger.warning(f"No products found in collection {collection_id}")
            return {
                "collection_id": collection_id,
                "success": True,
                "total": 0,
                "updated": 0,
                "skipped": 0,
                "failed": 0,
                "results": []
            }

        # Prepare updates
        updates = []
        results = []

        for product in products:
            # Get detailed product information
            product_details = self.get_product_details(product.id)
            if not product_details:
                results.append(TypeSyncResult(
                    product_id=product.id,
                    product_title=product.title,
                    collection_id=collection_id,
                    success=False,
                    old_type=None,
                    new_type="",
                    error="Failed to fetch product details"
                ))
                continue

            # Determine target type
            # Tags are already handled in get_product_details as a list
            product_tags = product_details.get("tags", [])
            
            target_type = self.determine_product_type(collection_id, product_tags)
            if not target_type:
                results.append(TypeSyncResult(
                    product_id=product.id,
                    product_title=product.title,
                    collection_id=collection_id,
                    success=False,
                    old_type=product_details.get("type"),
                    new_type="",
                    error="Could not determine target type"
                ))
                continue

            current_type = product_details.get("type")
            
            # Check if update is needed
            if current_type == target_type:
                results.append(TypeSyncResult(
                    product_id=product.id,
                    product_title=product.title,
                    collection_id=collection_id,
                    success=True,
                    old_type=current_type,
                    new_type=target_type,
                    skipped=True,
                    reason=f"Type already set to '{target_type}'"
                ))
                continue

            # Prepare update
            updates.append({
                "product_id": product.id,
                "product_gid": product_details.get("gid"),
                "product_title": product.title,
                "current_type": current_type,
                "target_type": target_type,
                "tags": product_tags
            })

        logger.info(f"📝 Products to update: {len(updates)}")
        logger.info(f"⏭️  Products to skip: {len([r for r in results if r.skipped])}")

        if dry_run:
            logger.info("🔍 DRY RUN - No changes will be made")
            for update in updates:
                logger.info(
                    f"  Would update: {update['product_title']} "
                    f"({update['current_type']} → {update['target_type']})"
                )
            
            return {
                "collection_id": collection_id,
                "collection_name": collection_name,
                "total": len(products),
                "to_update": len(updates),
                "skipped": len([r for r in results if r.skipped]),
                "failed": len([r for r in results if not r.success and not r.skipped]),
                "dry_run": True,
                "results": results
            }

        # Execute updates concurrently
        if updates:
            logger.info(f"🔄 Updating {len(updates)} products...")

            updated_count = 0
            failed_count = 0

            with ThreadPoolExecutor(max_workers=self.MAX_CONCURRENT_UPDATES) as executor:
                future_to_update = {
                    executor.submit(
                        self.update_product_type,
                        update["product_id"],
                        update["product_gid"],
                        update["target_type"]
                    ): update
                    for update in updates
                }

                for future in as_completed(future_to_update):
                    update = future_to_update[future]
                    success = future.result()

                    if success:
                        updated_count += 1
                        results.append(TypeSyncResult(
                            product_id=update["product_id"],
                            product_title=update["product_title"],
                            collection_id=collection_id,
                            success=True,
                            old_type=update["current_type"],
                            new_type=update["target_type"]
                        ))
                        logger.info(
                            f"✅ [{updated_count}/{len(updates)}] Updated: {update['product_title']} "
                            f"({update['current_type']} → {update['target_type']})"
                        )
                    else:
                        failed_count += 1
                        results.append(TypeSyncResult(
                            product_id=update["product_id"],
                            product_title=update["product_title"],
                            collection_id=collection_id,
                            success=False,
                            old_type=update["current_type"],
                            new_type=update["target_type"],
                            error="Update failed"
                        ))
                        logger.error(
                            f"❌ [{updated_count + failed_count}/{len(updates)}] Failed: "
                            f"{update['product_title']}"
                        )

                    # Small delay to avoid rate limiting
                    time.sleep(0.1)
        else:
            updated_count = 0
            failed_count = 0

        # Summary
        skipped_count = len([r for r in results if r.skipped])
        logger.info("\n" + "="*60)
        logger.info(f"📊 COLLECTION SUMMARY: {collection_name}")
        logger.info("="*60)
        logger.info(f"Total products: {len(products)}")
        logger.info(f"✅ Updated: {updated_count}")
        logger.info(f"⏭️  Skipped: {skipped_count}")
        logger.info(f"❌ Failed: {failed_count}")
        logger.info("="*60)

        return {
            "collection_id": collection_id,
            "collection_name": collection_name,
            "total": len(products),
            "updated": updated_count,
            "skipped": skipped_count,
            "failed": failed_count,
            "results": results
        }

    def sync_all_collections(self, dry_run: bool = False) -> Dict[str, Any]:
        """
        Process all configured collections.

        Args:
            dry_run: If True, only analyze without making changes

        Returns:
            Dictionary with overall sync results
        """
        logger.info("="*60)
        logger.info("🚀 STARTING PRODUCT TYPE SYNC")
        logger.info("="*60)
        logger.info(f"Collections to process: {len(self.COLLECTIONS)}")
        logger.info(f"Dry run: {dry_run}")
        logger.info("="*60)

        all_results = {}
        total_updated = 0
        total_skipped = 0
        total_failed = 0
        total_products = 0

        for collection_id in self.COLLECTIONS.keys():
            result = self.process_collection(collection_id, dry_run=dry_run)
            all_results[collection_id] = result

            total_updated += result.get("updated", 0)
            total_skipped += result.get("skipped", 0)
            total_failed += result.get("failed", 0)
            total_products += result.get("total", 0)

            # Small delay between collections
            time.sleep(0.5)

        # Overall summary
        logger.info("\n" + "="*60)
        logger.info("📊 OVERALL SYNC SUMMARY")
        logger.info("="*60)
        logger.info(f"Total products processed: {total_products}")
        logger.info(f"✅ Total updated: {total_updated}")
        logger.info(f"⏭️  Total skipped: {total_skipped}")
        logger.info(f"❌ Total failed: {total_failed}")
        logger.info("="*60)

        return {
            "total_products": total_products,
            "total_updated": total_updated,
            "total_skipped": total_skipped,
            "total_failed": total_failed,
            "collections": all_results,
            "dry_run": dry_run
        }

    def get_products_with_empty_type(self) -> List[Dict[str, Any]]:
        """
        Get all ACTIVE products with empty or null product type.

        Returns:
            List of active products with empty type, each containing id, title, handle, and productType
        """
        logger.info("🔍 Fetching all ACTIVE products with empty product type...")
        logger.info("="*60)

        query = """
        query getProducts($cursor: String, $query: String) {
          products(first: 50, after: $cursor, query: $query) {
            pageInfo {
              hasNextPage
              endCursor
            }
            edges {
              cursor
              node {
                id
                legacyResourceId
                title
                handle
                productType
                status
              }
            }
          }
        }
        """

        products_with_empty_type = []
        cursor = None
        has_next_page = True
        total_checked = 0
        
        # Query only ACTIVE products
        query_string = "status:active"

        while has_next_page:
            variables = {
                "query": query_string
            }
            if cursor:
                variables["cursor"] = cursor
                
            result = self.shopify_client.execute_graphql(query, variables)

            if not result or "data" not in result:
                logger.error("Failed to fetch products")
                break

            if "errors" in result:
                error_messages = [err.get("message", "Unknown error") for err in result["errors"]]
                logger.error(f"GraphQL errors: {', '.join(error_messages)}")
                break

            edges = result["data"]["products"]["edges"]
            page_info = result["data"]["products"]["pageInfo"]

            for edge in edges:
                node = edge["node"]
                total_checked += 1
                
                # Double-check status (should be ACTIVE from query, but verify)
                status = node.get("status")
                if status != "ACTIVE":
                    continue  # Skip non-active products
                
                product_type = node.get("productType")
                
                # Check if product type is empty, null, or whitespace
                if not product_type or (isinstance(product_type, str) and product_type.strip() == ""):
                    products_with_empty_type.append({
                        "id": node.get("legacyResourceId"),
                        "gid": node.get("id"),
                        "title": node.get("title"),
                        "handle": node.get("handle"),
                        "productType": product_type or "",
                        "status": status
                    })

            # Update pagination
            has_next_page = page_info.get("hasNextPage", False)
            if has_next_page and edges:
                cursor = edges[-1]["cursor"]

            # Progress update
            if total_checked % 250 == 0:
                logger.info(f"Checked {total_checked} products, found {len(products_with_empty_type)} with empty type...")

            # Small delay to avoid rate limiting
            time.sleep(0.1)

        logger.info("="*60)
        logger.info(f"✅ Checked {total_checked} ACTIVE products total")
        logger.info(f"📊 Found {len(products_with_empty_type)} ACTIVE products with empty product type")
        logger.info("="*60)

        return products_with_empty_type

    def export_products_with_empty_type_to_excel(
        self,
        products: List[Dict[str, Any]],
        output_dir: Optional[str] = None
    ) -> str:
        """
        Export products with empty type to Excel file.

        Args:
            products: List of products with empty type
            output_dir: Optional output directory (default: project_root/data)

        Returns:
            Path to the created Excel file
        """
        if not products:
            logger.warning("No products to export")
            return ""

        # Determine output directory
        if output_dir:
            output_path = Path(output_dir)
        else:
            # Default to project_root/data
            project_root = Path(__file__).parent.parent.parent
            output_path = project_root / "data"
        
        output_path.mkdir(parents=True, exist_ok=True)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"products_empty_type_{timestamp}.xlsx"
        file_path = output_path / filename

        logger.info(f"📝 Creating Excel file: {file_path}")

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products with Empty Type"

        # Define headers
        headers = [
            "Product ID",
            "Product Title",
            "Handle",
            "Product Type",
            "Status"
        ]

        # Style for header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data
        for row_idx, product in enumerate(products, 2):
            ws.cell(row=row_idx, column=1, value=product.get("id", ""))
            ws.cell(row=row_idx, column=2, value=product.get("title", ""))
            ws.cell(row=row_idx, column=3, value=product.get("handle", ""))
            ws.cell(row=row_idx, column=4, value=product.get("productType", "") or "")
            ws.cell(row=row_idx, column=5, value=product.get("status", "ACTIVE"))

        # Auto-adjust column widths
        column_widths = {
            "A": 15,  # Product ID
            "B": 50,  # Product Title
            "C": 30,  # Handle
            "D": 20,  # Product Type
            "E": 12   # Status
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Add summary row
        summary_row = len(products) + 3
        ws.cell(row=summary_row, column=1, value="Total:").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=len(products)).font = Font(bold=True)
        ws.cell(row=summary_row, column=3, value="ACTIVE products with empty type")

        # Save workbook
        wb.save(file_path)
        logger.info(f"✅ Excel file created successfully: {file_path}")

        return str(file_path)
