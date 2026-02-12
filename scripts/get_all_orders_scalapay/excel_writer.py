#!/usr/bin/env python3
"""
Excel Writer for Scalapay Orders

Generates Excel reports with Scalapay order data.
"""

import logging
from pathlib import Path
from typing import List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scripts.get_all_orders_scalapay.models import ScalapayOrder, ScalapayOrderResult

logger = logging.getLogger(__name__)


class ScalapayExcelWriter:
    """Excel writer for Scalapay order reports"""

    # Column definitions
    COLUMNS = [
        ("Order Name", 15),
        ("Created At", 18),
        ("Customer Email", 35),
        ("Customer Name", 25),
        ("Fulfillment Status", 18),
        ("Financial Status", 16),
        ("Total Price", 12),
        ("Subtotal", 12),
        ("Shipping", 12),
        ("Refunded", 12),
        ("Currency", 10),
        ("Shipping City", 18),
        ("Shipping Country", 16),
        ("Shipping Address", 50),
        ("Line Items", 80),
        ("Payment Gateway", 20),
    ]

    # Styles
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    UNFULFILLED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    UNPAID_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def __init__(self, output_path: str):
        """
        Initialize Excel writer

        Args:
            output_path: Path for output Excel file
        """
        self.output_path = Path(output_path)
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Scalapay Orders"

    def write_orders(self, result: ScalapayOrderResult) -> bool:
        """
        Write orders to Excel file

        Args:
            result: ScalapayOrderResult containing orders and statistics

        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Writing {len(result.orders)} orders to Excel...")

            # Write summary sheet
            self._write_summary(result)

            # Write data sheet
            self._write_data_sheet(result.orders)

            # Ensure output directory exists
            self.output_path.parent.mkdir(parents=True, exist_ok=True)

            # Save workbook
            self.workbook.save(self.output_path)
            logger.info(f"✅ Excel file saved to: {self.output_path}")

            return True

        except Exception as e:
            logger.error(f"Error writing Excel file: {e}")
            return False

    def _write_summary(self, result: ScalapayOrderResult):
        """Write summary sheet"""
        # Create summary sheet
        summary_sheet = self.workbook.create_sheet("Summary", 0)

        # Title
        summary_sheet["A1"] = "Scalapay Orders Report"
        summary_sheet["A1"].font = Font(bold=True, size=16)
        summary_sheet.merge_cells("A1:C1")

        # Report date
        summary_sheet["A3"] = "Report Generated:"
        summary_sheet["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Statistics
        summary_sheet["A5"] = "Statistics"
        summary_sheet["A5"].font = Font(bold=True, size=12)

        stats = [
            ("Total Orders Scanned", result.total_orders_scanned),
            ("Scalapay Orders Found", result.scalapay_orders_found),
            ("Percentage", f"{result.success_rate:.2f}%"),
            ("Execution Time", f"{result.execution_time_seconds:.2f} seconds"),
        ]

        for i, (label, value) in enumerate(stats, start=6):
            summary_sheet[f"A{i}"] = label
            summary_sheet[f"B{i}"] = value

        # Fulfillment status breakdown
        if result.orders:
            summary_sheet["A12"] = "Fulfillment Status Breakdown"
            summary_sheet["A12"].font = Font(bold=True, size=12)

            status_counts = {}
            for order in result.orders:
                status = order.fulfillment_status_display
                status_counts[status] = status_counts.get(status, 0) + 1

            row = 13
            for status, count in sorted(status_counts.items()):
                summary_sheet[f"A{row}"] = status
                summary_sheet[f"B{row}"] = count
                row += 1

        # Financial status breakdown
        if result.orders:
            summary_sheet[f"A{row + 1}"] = "Financial Status Breakdown"
            summary_sheet[f"A{row + 1}"].font = Font(bold=True, size=12)

            financial_counts = {}
            for order in result.orders:
                status = order.financial_status_display
                financial_counts[status] = financial_counts.get(status, 0) + 1

            row = row + 2
            for status, count in sorted(financial_counts.items()):
                summary_sheet[f"A{row}"] = status
                summary_sheet[f"B{row}"] = count
                row += 1

        # Adjust column widths
        summary_sheet.column_dimensions["A"].width = 30
        summary_sheet.column_dimensions["B"].width = 25

    def _write_data_sheet(self, orders: List[ScalapayOrder]):
        """Write data sheet with order details"""
        ws = self.worksheet

        # Write headers
        for col_idx, (col_name, col_width) in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Set column width
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Write data rows
        for row_idx, order in enumerate(orders, start=2):
            col = 1

            # Order Name
            ws.cell(row=row_idx, column=col, value=order.order_name)
            col += 1

            # Created At
            if order.created_at:
                ws.cell(row=row_idx, column=col, value=order.created_at.strftime("%Y-%m-%d %H:%M"))
            else:
                ws.cell(row=row_idx, column=col, value="")
            col += 1

            # Customer Email
            ws.cell(row=row_idx, column=col, value=order.customer_email or "")
            col += 1

            # Customer Name
            ws.cell(row=row_idx, column=col, value=order.customer_full_name)
            col += 1

            # Fulfillment Status
            fulfillment_cell = ws.cell(row=row_idx, column=col, value=order.fulfillment_status_display)
            if order.fulfillment_status_display == "UNFULFILLED":
                fulfillment_cell.fill = self.UNFULFILLED_FILL
            col += 1

            # Financial Status
            financial_cell = ws.cell(row=row_idx, column=col, value=order.financial_status_display)
            if order.financial_status_display not in ["PAID", "AUTHORIZED"]:
                financial_cell.fill = self.UNPAID_FILL
            col += 1

            # Total Price
            ws.cell(row=row_idx, column=col, value=order.total_price or "")
            col += 1

            # Subtotal
            ws.cell(row=row_idx, column=col, value=order.subtotal_price or "")
            col += 1

            # Shipping
            ws.cell(row=row_idx, column=col, value=order.shipping_price or "")
            col += 1

            # Refunded
            refunded_cell = ws.cell(row=row_idx, column=col, value=order.total_refunded or "")
            # Highlight if there's a refund
            if order.total_refunded and float(order.total_refunded) > 0:
                refunded_cell.fill = self.UNPAID_FILL
            col += 1

            # Currency
            ws.cell(row=row_idx, column=col, value=order.currency or "")
            col += 1

            # Shipping City
            ws.cell(row=row_idx, column=col, value=order.shipping_city or "")
            col += 1

            # Shipping Country
            ws.cell(row=row_idx, column=col, value=order.shipping_country or "")
            col += 1

            # Shipping Address
            ws.cell(row=row_idx, column=col, value=order.shipping_address or "")
            col += 1

            # Line Items
            ws.cell(row=row_idx, column=col, value=order.line_items or "")
            col += 1

            # Payment Gateway
            ws.cell(row=row_idx, column=col, value=order.payment_gateway)

        # Add auto-filter
        if orders:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(self.COLUMNS))}{len(orders) + 1}"

        logger.info(f"Written {len(orders)} rows to data sheet")

    def close(self):
        """Close workbook"""
        try:
            self.workbook.close()
        except Exception:
            pass
